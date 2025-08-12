import re
import yaml
import argparse
import os
import pandas as pd
from natsort import natsorted
import openpyxl
import openpyxl.utils
import shutil
from template import template

# Stellt sicher, dass der Typ klein- und großgeschrieben werden kann
def case_insensitive(choices):
    lower_choices = [c.lower() for c in choices]
    def check_choice(arg):
        if arg.lower() in lower_choices:
            return choices[lower_choices.index(arg.lower())]
        else:
            raise argparse.ArgumentTypeError(
                f"invalid choice: '{arg}' (choose from {', '.join(choices)})"
            )
    return check_choice

# Erstellt eine neue Markdown-Datei nach Template, falls sie nicht existiert                
def new_file_with_template(directory):
    if not os.path.isdir(directory):
        print(f"{directory} ist kein gültiger Ordner, bitte Zielordner angeben.")
        return
    filepath = f"{directory}/neue_richtlinie.md"
    os.makedirs(directory, exist_ok=True)

    if not os.path.exists(filepath):
        with open(filepath, "w", encoding="utf-8") as file:
            file.write(template)
        print(f"Markdown-Datei '{filepath}' wurde erstellt.")
    else:
        print(f"Die Datei '{filepath}' existiert bereits. Vorgang abgebrochen")

# Vereinheitlicht IDs, falls eindeutige IDs verwendet werden
def id_unify(id):
    id = id.upper().strip()
    id = id.replace("-", ".")
    if id[-1].isalpha():
        id = id[:-1]
    if 'A' in id and '.A' not in id:
        last_a_index = id.rfind('A')
        if last_a_index != -1:
            id = id[:last_a_index] + '.' + id[last_a_index:]
    return id

def get_bausteine_from_metadata(content, file_path, valid_baustein_ids, verbose):
    baustein_references = set()
    metadata_match = re.match(r"^---\n(.*?)\n---\n", content, re.DOTALL)
    if not metadata_match and verbose:
        print(f"\nKeine Metadaten in {file_path} gefunden.")
        return

    metadata_block = metadata_match.group(1)
    metadata = yaml.safe_load(metadata_block)
    bsi_references = metadata.get("refs", [{}])[0].get("bsi")

    if not bsi_references and verbose:
        print(f"\nKeine BSI IT-Grundschutz Referenzen in den Metadaten von {file_path} gefunden.")
        return

    for bsi_reference in bsi_references:
        main_reference = bsi_reference.upper()
        main_reference = main_reference.split('.A')[0]

        main_reference = main_reference.replace("-", ".")

        if main_reference in valid_baustein_ids:
            baustein_references.add(main_reference)
        elif verbose:
            print(f"WARNUNG: Ungültige Bausteinreferenz in Metadaten: {main_reference}")
    return baustein_references

def get_reqs_from_comments(content, valid_anforderung_ids, verbose):
    comment_pattern = re.compile(r"<!--\s*(.*?)\s*-->", re.DOTALL)

    comment_contents = comment_pattern.findall(content)

    erfuellt = []
    teilweise = []
    entbehrlich = []

    allowed_statuses = {"erfüllt", "umgesetzt", "teilweise", "entbehrlich"}
    id_format_pattern = re.compile(r"^(?:orp|ops|sys|app|con|isms|der|ind|net|inf)\.\d+(?:\.\d+)*\.A\d+$", re.IGNORECASE)

    # Hilfsfunktion, um zu prüfen, ob ein String IDs enthält
    def contains_valid_id(text):
        potential_ids = re.split(r'[\s,;]+', text)
        for pid in potential_ids:
            pid = id_unify(pid)
            if id_format_pattern.match(pid):
                return True
        return False

    for item in comment_contents:
        cleaned_item = item.strip()

        parts = cleaned_item.rsplit(':', 1)
        status = "erfüllt"
        ids_part = cleaned_item

        comment_has_valid_ids = contains_valid_id(parts[0].strip())

        if len(parts) == 2:
            potential_ids_part = parts[0].strip()
            potential_status = parts[1].strip().lower()

            if potential_status in allowed_statuses:
                status = potential_status
                ids_part = potential_ids_part
            else:
                if comment_has_valid_ids:
                    raise ValueError(
                        f"Ungültiger Umsetzungsstatus '{parts[1].strip()}' in Kommentar: '{item}'. "
                        f"Erlaubt sind: {', '.join(allowed_statuses)}."
                    )

        found_ids_in_comment = re.split(r'[\s,;]+', ids_part)

        for single_id in found_ids_in_comment:
            if not single_id:
                continue

            unified_id = id_unify(single_id)

            if id_format_pattern.match(unified_id):
                if unified_id in valid_anforderung_ids:
                    if status in {"erfüllt", "umgesetzt"}:
                        erfuellt.append(unified_id)
                    elif status == "teilweise":
                        teilweise.append(unified_id)
                    elif status == "entbehrlich":
                        entbehrlich.append(unified_id)
                elif verbose:
                    print(f"WARNUNG: '{single_id}' (verarbeitet als '{unified_id}') ist keine gültige ID (evtl. entfallen).")
            else: break

        if not comment_has_valid_ids and verbose:
            print(f"Info: Kommentar ignoriert: '{item}'")
    return erfuellt, teilweise, entbehrlich


# Überprüft den Umsetzungsstatus von Anforderungen in Richtlinien
def check(path, typ, show_status, inhalt, verbose):
    files = set()
    if os.path.isfile(path):
        files.add(path)
    elif os.path.isdir(path):
        for file in os.listdir(path):
            file_path = os.path.join(path, file)
            if os.path.isfile(file_path) and file_path.endswith(('.md', '.markdown')):
                files.add(file_path)
    else:
        print("Kein gültiger Pfad, Vorgang wird abgebrochen.")
        return

    try:
        script_dir = os.path.dirname(__file__)
        json_path = os.path.join(script_dir, "IT-Grundschutz.json")
        it_gs = pd.read_json(json_path, orient="records")
    except FileNotFoundError:
        print("\nFehler: Die Datei 'IT-Grundschutz.json' wurde nicht im Skriptverzeichnis gefunden.")
        return

    it_gs['Baustein_ID'] = it_gs['ID-Anforderung'].str.split('.A', expand=True)[0]
    it_gs['ID_unified'] = it_gs['ID-Anforderung'].apply(id_unify)
    valid_baustein_ids = set(it_gs['Baustein_ID'].unique())
    valid_anforderung_ids = set(it_gs['ID_unified'].unique())


    for file_path in files:
        with open(file_path, 'r') as file:
            content = file.read()

        baustein_references = get_bausteine_from_metadata(content, file_path, valid_baustein_ids, verbose)
        if not baustein_references:
            continue

        print(f"\n\n========================================================")
        print(f"   Analyse für Datei: {os.path.basename(file_path)}")
        print(f"========================================================")

        erfuellt, teilweise, entbehrlich = get_reqs_from_comments(content, valid_anforderung_ids, verbose)

        relevante_anforderungen = it_gs[it_gs['Baustein_ID'].isin(baustein_references)].copy()

        alle_gefundenen_ids = set(erfuellt + teilweise + entbehrlich)
        relevante_ids = set(relevante_anforderungen['ID_unified'])
        falsch_dokumentierte_ids = alle_gefundenen_ids - relevante_ids

        if falsch_dokumentierte_ids and verbose:
            for anforderung_id in list(falsch_dokumentierte_ids):
                baustein_id = anforderung_id.split('.A')[0]
                print(f"WARNUNG: '{anforderung_id}' wird genannt, aber der Baustein '{baustein_id}' ist nicht in den Metadaten der Datei.")

        if typ.lower() == 'basis':
            relevante_anforderungen = relevante_anforderungen[relevante_anforderungen['Typ'] == 'Basis']
        elif typ.lower() == 'standard':
            relevante_anforderungen = relevante_anforderungen[relevante_anforderungen['Typ'].isin(['Basis', 'Standard'])]
        # Bei Hoch werden alle beibehalten

        alle_relevanten_ids = set(relevante_anforderungen['ID_unified'])

        tmp_erfuellt = [id_ for id_ in erfuellt if id_ in alle_relevanten_ids]
        tmp_teilweise = [id_ for id_ in teilweise if id_ in alle_relevanten_ids]
        tmp_entbehrlich = [id_ for id_ in entbehrlich if id_ in alle_relevanten_ids]

        rel_erfuellt = natsorted(tmp_erfuellt)
        rel_teilweise = natsorted(tmp_teilweise)
        rel_entbehrlich = natsorted(tmp_entbehrlich)
        nicht_umgesetzte_ids = natsorted(list(alle_relevanten_ids - set(rel_erfuellt + rel_teilweise + rel_entbehrlich)))

        print("\n--- Übersicht ---")
        print(f"Schutzbedarfstyp: {typ.capitalize()}")
        print(f"Umgesetzt: {len(rel_erfuellt)}")
        print(f"Teilweise umgesetzt: {len(rel_teilweise)}")
        print(f"Entbehrlich: {len(rel_entbehrlich)}")
        print(f"Nicht umgesetzt / Nicht dokumentiert: {len(nicht_umgesetzte_ids)}")
        if show_status:
            print()

        # Hilfsfunktion für formatierte Ausgabe
        def print_section(title, ids, df, status, inhalt):
            if not status:
                print(f"\n--- {title} ({len(ids)}) ---")
                if not ids:
                    print("Keine")
                    return

            status_map = {
                "Umgesetzt": "(umgesetzt)",
                "Teilweise umgesetzt": "(teilweise)",
                "Entbehrlich": "(entbehrlich)"
            }
            status_text = status_map.get(title, "(offen)")

            grouped_items = {}
            for anforderung_id in ids:
                details = df[df['ID_unified'] == anforderung_id]
                if not details.empty:
                    for index, detail in details.iterrows():
                        key = (detail['ID_unified'], detail['Titel'], detail['Baustein'], detail['Typ'])
                        if key not in grouped_items:
                            grouped_items[key] = []
                        grouped_items[key].append(detail['Inhalt'])

            sorted_keys = natsorted(grouped_items.keys())

            for id_title_key in sorted_keys:
                anforderung_id, title, baustein, typ_detail = id_title_key
                contents = grouped_items[id_title_key]

                if not inhalt and not status:
                    print(f"{anforderung_id} - {title}")

                if not inhalt and status:
                    print(f"{anforderung_id} - {title} {status_text}")

                if inhalt and not status:
                    unique_contents = []
                    [unique_contents.append(x) for x in contents if x not in unique_contents]
                    print(f"\n\n{anforderung_id} - {title}")
                    print(f"Baustein: {baustein}\nTyp: {typ_detail}")
                    print(f"Inhalt: {' '.join(unique_contents)}")

                if inhalt and status:
                    unique_contents = []
                    [unique_contents.append(x) for x in contents if x not in unique_contents]
                    print(f"\n\n{anforderung_id} - {title} {status_text}")
                    print(f"Baustein: {baustein}\nTyp: {typ_detail}")
                    print(f"Inhalt: {' '.join(unique_contents)}")

        print_section("Umgesetzt", rel_erfuellt, relevante_anforderungen, show_status, inhalt)
        print_section("Teilweise umgesetzt", rel_teilweise, relevante_anforderungen, show_status, inhalt)
        print_section("Entbehrlich", rel_entbehrlich, relevante_anforderungen, show_status, inhalt)
        print_section("Nicht umgesetzt / Nicht dokumentiert", nicht_umgesetzte_ids, relevante_anforderungen, show_status, inhalt)

# Gibt den Inhalt einer gewünschten Anforderung aus
def explain(string):
    id = id_unify(string)
    try:
        script_dir = os.path.dirname(__file__)
        json_path = os.path.join(script_dir, "IT-Grundschutz.json")
        it_gs = pd.read_json(json_path, orient="records")
    except FileNotFoundError:
        print("\nFehler: Die Datei 'IT-Grundschutz.json' wurde nicht im Skriptverzeichnis gefunden.")
        return

    it_gs['ID_unified'] = it_gs['ID-Anforderung'].apply(id_unify)

    if 'ID-Anforderung' in it_gs.columns:
        it_gs['Baustein_ID'] = it_gs['ID-Anforderung'].str.split('.A', expand=True)[0]

    results = it_gs[it_gs['ID_unified'] == id]

    if results.empty:
        print(f"Anforderung '{string}' wurde nicht im IT-Grundschutz-Kompendium gefunden (evtl. entfallen).")
        return

    first_row = results.iloc[0]

    content_parts = results['Inhalt'].dropna().tolist()
    unique_contents = []
    [unique_contents.append(x) for x in content_parts if x not in unique_contents]
    full_content = ' '.join(unique_contents)

    print(f"\n--- Details für Anforderung: {id} ---")
    print(f"Titel: {first_row.get('Titel')}")
    print(f"Baustein: {first_row.get('Baustein')}")
    print(f"Typ: {first_row.get('Typ')}")
    print(f"Inhalt: {full_content}")

# Füllt Checklisten aus Richtlinien aus
def fill_checklists(path, verbose):
    files = set()
    if os.path.isfile(path):
        files.add(path)
    elif os.path.isdir(path):
        for file in os.listdir(path):
            file_path = os.path.join(path, file)
            if os.path.isfile(file_path) and file_path.endswith(('.md', '.markdown')):
                files.add(file_path)
    else:
        print("Kein gültiger Pfad, Vorgang wird abgebrochen.")
        return

    print(f"Pfad zu Richtlinie(n): {path}")
    while True:
        checklist_path = input("\nPfad zum Ordner mit den Checklisten: ").strip().strip('"')
        if os.path.isdir(checklist_path):
            break
        else:
            print("Kein gültiger Pfad zu einem Ordner")

    try:
        script_dir = os.path.dirname(__file__)
        json_path = os.path.join(script_dir, "IT-Grundschutz.json")
        it_gs = pd.read_json(json_path, orient="records")
    except FileNotFoundError:
        print("\nFehler: Die Datei 'IT-Grundschutz.json' wurde nicht im Skriptverzeichnis gefunden.")
        return

    it_gs['Baustein_ID'] = it_gs['ID-Anforderung'].str.split('.A', expand=True)[0]
    it_gs['ID_unified'] = it_gs['ID-Anforderung'].apply(id_unify)
    valid_baustein_ids = set(it_gs['Baustein_ID'].unique())
    valid_anforderung_ids = set(it_gs['ID_unified'].unique())

    for file_path in files:
        execute_checklist_filling(file_path, checklist_path, it_gs, valid_baustein_ids, valid_anforderung_ids, verbose)
    print(f"Checklisten gefüllt und gespeichert")

# Führt die Bearbeitung der Checklisten für jede Datei aus
def execute_checklist_filling(file_path, checklist_path, it_gs, valid_baustein_ids, valid_anforderung_ids, verbose):
    with open(file_path, 'r') as file:
        content = file.read()

    baustein_references = get_bausteine_from_metadata(content, file_path, valid_baustein_ids, verbose)
    if not baustein_references:
        return

    checklist_files_to_process = []
    for baustein_id in baustein_references:
        potential_checklist = os.path.join(checklist_path, f"Checkliste_{baustein_id}.xlsx")
        if os.path.exists(potential_checklist):
            checklist_files_to_process.append(potential_checklist)
        else:
            print(f"WARNUNG: Checkliste für Baustein '{baustein_id}' nicht gefunden unter '{potential_checklist}'. Wird übersprungen.")

    if not checklist_files_to_process:
        return

    erfuellt, teilweise, entbehrlich = get_reqs_from_comments(content, valid_anforderung_ids, verbose)

    relevante_anforderungen = it_gs[it_gs['Baustein_ID'].isin(baustein_references)].copy()

    alle_relevanten_ids = set(relevante_anforderungen['ID_unified'])
    rel_erfuellt = [id_ for id_ in erfuellt if id_ in alle_relevanten_ids]
    rel_teilweise = [id_ for id_ in teilweise if id_ in alle_relevanten_ids]
    rel_entbehrlich = [id_ for id_ in entbehrlich if id_ in alle_relevanten_ids]
    nicht_umgesetzte_ids = list(alle_relevanten_ids - set(rel_erfuellt + rel_teilweise + rel_entbehrlich))

    output_dir = os.path.join(checklist_path, "Aus Richtlinie generiert")
    os.makedirs(output_dir, exist_ok=True)

    for source_checklist_path in checklist_files_to_process:
        base_name = os.path.basename(source_checklist_path)
        dest_checklist_path = os.path.join(output_dir, base_name)

        shutil.copy(source_checklist_path, dest_checklist_path)

        workbook = openpyxl.load_workbook(dest_checklist_path)
        sheet = workbook.active

        headers = {cell.value: cell.column for cell in sheet[5]}
        id_col = headers.get("ID-Anforderung")
        umsetzung_col = headers.get("Umsetzung")
        entbehrlich_col = headers.get("Entbehrlich")

        for row in range(6, sheet.max_row + 1):
            req_id_cell = sheet.cell(row=row, column=id_col)
            if req_id_cell.value:
                unified_id = id_unify(req_id_cell.value)

                umsetzung_cell = sheet.cell(row=row, column=umsetzung_col)
                current_value = umsetzung_cell.value

                if unified_id in rel_erfuellt:
                    umsetzung_cell.value = "Ja"
                elif unified_id in rel_teilweise:
                    if current_value is None or current_value == "Nein":
                        umsetzung_cell.value = "Teilweise"
                elif unified_id in nicht_umgesetzte_ids:
                    if current_value is None:
                        umsetzung_cell.value = "Nein"
                if unified_id in rel_entbehrlich:
                    sheet.cell(row=row, column=entbehrlich_col).value = "Ja"

        workbook.save(dest_checklist_path)

# Hauptprogramm zur Verarbeitung von Kommandozeilenargumenten
def main():
    parser = argparse.ArgumentParser(description='Hilft bei der Prüfung von Markdown-Richtlinien.')
    type_choices = ['Basis', 'Standard', 'Hoch']
    parser.add_argument('file_dir_or_string', help='Pfad zu Datei, Verzeichnis oder String')
    parser.add_argument("--check", action="store_true", help="Prüft die Anforderungen von Richtlinien auf Vollständigkeit")
    parser.add_argument('--typ', type=case_insensitive(type_choices), choices=type_choices, default='Standard', help='(Mit --check) Gibt den Typ bei der Anforderungsprüfung an (Default: Standard})')
    parser.add_argument("--status", action="store_true", help="(Mit --check) Zeigt zusätzlich in jeder Zeile den Umsetzungsstatus an")
    parser.add_argument("--details", action="store_true", help="(Mit --check) Zeigt zusätzlich Baustein und Inhalt an")
    parser.add_argument("--explain", action="store_true", help="Zeigt Details zu einer bestimmten Anforderung an")
    parser.add_argument("--new", action="store_true", help="Erstellt eine neue Richtlinien nach Template im übergebenen Ordner")
    parser.add_argument("--fill-checklists", action="store_true", help="Füllt automatisch die Checklisten basierend auf der Richtlinie aus")
    parser.add_argument("-v", "--verbose", action="store_true", help="Zeigt Infos und Warnungen während der Verarbeitung an")

    args = parser.parse_args()

    if args.check:
        check(args.file_dir_or_string, args.typ, args.status, args.details, args.verbose)
        return

    elif args.explain:
        explain(args.file_dir_or_string)
        return

    elif args.new:
        new_file_with_template(args.file_dir_or_string)
        return

    elif args.fill_checklists:
        fill_checklists(args.file_dir_or_string, args.verbose)
        return

    else: print("Kein Argument übergeben")

if __name__ == '__main__':
    main()